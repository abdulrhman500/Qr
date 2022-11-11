from genericpath import isfile
from PIL import Image
from openpyxl import load_workbook
import pandas as pd
import qrcode
import os.path
import warnings
import xlsxwriter
import stat

warnings.filterwarnings('ignore')

File_Path = os.path.abspath('./example.xlsx')
Images_Path = "./Images/"
Image_Title = 'Person no. '
Logo_Path = os.path.abspath("./Images/logo.png")
Sheet_Name='Sheet1'

class QrGenerator:

    def __init__(self, path, logo, save, title,sheetname):
        exist = isfile(path)
        if not exist:
            print("[ERROR]: File does not exist. Path entered is : " + path)
            exit()
        self.logo = logo
        self.path = path
        self.save = save
        self.title = title
        self.links_list = []
        self.sheetname=sheetname
        self.read()
        self.qr_generator()
        self.write()

    def read(self):
        self.data = pd.read_excel(self.path)
        self.no_of_rows = len(self.data)
        self.no_of_columns = len(list(self.data))

    def get_data(self, i):
        data_list = str(self.data.loc[i]).splitlines()
        data = ''
        for element in range(len(data_list) - 1):
            data += (data_list[element] + '\n')
        return data

    def generate_img(self, qr):
        logo = Image.open(self.logo)
        basewidth = 100
        wpercent = (basewidth / float(logo.size[0]))
        hsize = int((float(logo.size[1]) * float(wpercent)))
        logo = logo.resize((basewidth, hsize), Image.ANTIALIAS)
        img = qr.make_image(fill_color='black', back_color="white").convert('RGB')
        pos = ((img.size[0] - logo.size[0]) // 2, (img.size[1] - logo.size[1]) // 2)
        img.paste(logo, pos)

        return img

    def qr_generator(self):
        for i in range(self.no_of_rows):
            QRcode = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H)

            data = self.get_data(i)

            QRcode.add_data(data)

            img = self.generate_img(QRcode)

            link = self.save + self.title + str(i) + ".png"

            self.links_list.append(link)

            img.save(link)

        return self.links_list

    def write(self):
        path = self.path
        qr_links = self.links_list
        os.chmod(path, stat.S_IWRITE)
        sheet_name = self.sheetname
        col = str(xlsxwriter.utility.xl_col_to_name(self.no_of_columns))
        wb = load_workbook(path)
        ws = wb[sheet_name]
        ws[col + str(1)] = 'QR_links'
        for i in range(2, self.no_of_rows + 2):
            cell = col + str(i)
            ws[cell].value = qr_links[i - 2]
        wb.save(path)


def main():
    QrGenerator(File_Path, Logo_Path, Images_Path, Image_Title, Sheet_Name)


if __name__ == "__main__":
    main()
